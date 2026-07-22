"""
Reads Expo City Dubai's existing ACC Naming Standard workbooks + folder structure
template + cascading metadata CSVs from OneDrive and compiles them into a single
machine-readable config/naming_rules.json used by the compliance auditor.

Source of truth (as found on disk 2026-07-22):
  RE&D - BIM_Internal/General/ACC Setup/ECD Setup ACC/
    Folder Structure/CXXXX - Contract Description - Folder Template/   <- required folder tree
    Naming Convention/<CODE>*/                                        <- per project-code naming standard

Re-run this any time the source workbooks change; it always regenerates
config/naming_rules.json from scratch (never hand-edit the generated file directly --
edit RESOLUTION_OVERRIDES / MANUAL_FIXUPS below instead, so re-runs stay reproducible).
"""

import glob
import json
import os
import re

import openpyxl

ONEDRIVE_ROOT = r"C:\Users\RE&DBIM\OneDrive - Expocitydubai"
NAMING_ROOT = os.path.join(ONEDRIVE_ROOT, r"RE&D - BIM_Internal\General\ACC Setup\ECD Setup ACC\Naming Convention")
FOLDER_TEMPLATE_ROOT = os.path.join(
    ONEDRIVE_ROOT,
    r"RE&D - BIM_Internal\General\ACC Setup\ECD Setup ACC\Folder Structure\CXXXX - Contract Description - Folder Template",
)
OUT_PATH = os.path.join(os.path.dirname(__file__), "..", "config", "naming_rules.json")

PROJECT_CODES = ["2EA", "2EM", "2RD", "2TD", "2DR", "2KR", "2LP", "2MR", "2TF", "2TR", "2TW"]

# Confirmed by direct inspection of ACC's own Account Admin > Naming Standards screen
# (Documentation/ACC_Folders_Attributes_Listing_InOrder.JPG). This is the hyphen-delimited
# file-name segment order; Revision/Status are frequently omitted on non-deliverable files
# (templates, starter-pack resources) so they're treated as optional trailing segments.
NAMING_SEGMENT_ORDER = ["Originator Organisation", "Type", "ProjectBuildingAsset", "Discipline", "Number", "Revision", "Status"]
REQUIRED_SEGMENTS = ["Originator Organisation", "Type", "ProjectBuildingAsset", "Discipline", "Number"]
OPTIONAL_TRAILING_SEGMENTS = ["Revision", "Status"]

# Full custom-attribute set shown in ACC Docs (Documentation/Attributes_Listing.JPG),
# excluding built-in (non-custom) fields Name/Description/Version/Review status.
# Treated as "should be populated" by default -- edit this list if some are meant optional.
METADATA_REQUIRED_FIELDS = [
    "Originator Organisation", "Type", "ProjectBuildingAsset", "Discipline", "Number",
    "Revision", "Status", "District", "Project", "Building / Asset Code",
    "Additional Description", "Stage", "Contract No",
]

# Sheet name -> attribute label used in NAMING_SEGMENT_ORDER / METADATA_REQUIRED_FIELDS.
# ("Building  Asset Code" has a double space in the source workbooks -- kept verbatim.)
SHEET_TO_LABEL = {
    "Originator": "Originator Organisation",
    "Type": "Type",
    "ProjectBuildingAsset": "ProjectBuildingAsset",
    "Discipline": "Discipline",
    "Number": "Number",
    "Revision": "Revision",
    "Status": "Status",
    "District": "District",
    "Project": "Project",
    "Building  Asset Code": "Building / Asset Code",
    "Additional Description": "Additional Description",
    "Stage": "Stage",
    "Contract": "Contract No",
}

EXCLUDE_PATH_MARKERS = ["superseed", "\\ss\\", "_ss\\", "_ss.xlsx"]


def is_excluded(path):
    lower = path.lower()
    return any(marker in lower for marker in EXCLUDE_PATH_MARKERS)


def read_project_code_from_workbook(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if "Project" not in wb.sheetnames:
        return None
    ws = wb["Project"]
    for row in ws.iter_rows(min_row=2, max_row=5, max_col=6):
        if row[3].value:
            code = str(row[3].value).split(" - ")[0].strip()
            return code
    return None


def find_canonical_workbook(code):
    """Pick the workbook that (a) isn't in a Superseed/SS path and (b) whose own
    'Project' sheet actually says this code -- among those, the most recently modified.
    Returns (path, status) where status is 'ok', 'multiple_candidates', or 'unresolved'."""
    candidates = []
    candidates += glob.glob(os.path.join(NAMING_ROOT, f"{code}_ACC_NamingStandard*.xlsx"))
    candidates += glob.glob(os.path.join(NAMING_ROOT, code, "*NamingStandard*.xlsx"))
    candidates = [c for c in candidates if not is_excluded(c)]

    matched = []
    mismatched = []
    for c in candidates:
        try:
            internal_code = read_project_code_from_workbook(c)
        except Exception as e:
            internal_code = f"ERROR:{e}"
        if internal_code == code:
            matched.append(c)
        else:
            mismatched.append((c, internal_code))

    if not matched:
        return None, "unresolved", mismatched
    matched.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    if len(matched) == 1:
        status = "ok"
    else:
        # NOTE: these files were all bulk-synced via OneDrive on the same day(s) --
        # mtimes differ by mere seconds and do NOT reflect true authorship recency.
        # Do not trust mtime to break the tie; surface it as ambiguous instead.
        status = "ambiguous_duplicates"
    return matched[0], status, [(m, "not picked -- unresolved duplicate, confirm manually") for m in matched[1:]] + [
        (m, f"internally labelled {ic}") for m, ic in mismatched
    ]


def extract_attribute_sheet(ws):
    """Returns list of {value, description} for a Drop-down list attribute sheet,
    or {char_type, length} for a Text field attribute sheet."""
    header = [c.value for c in ws[1]]
    if header[:3] == ["Name", "Description", "Type"] and len(header) >= 6 and header[3] == "Char type":
        row2 = [c.value for c in ws[2]] if ws.max_row >= 2 else []
        return {"kind": "text", "char_type": row2[3] if len(row2) > 3 else None, "length": row2[4] if len(row2) > 4 else None}
    values = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=6):
        val = row[3].value
        if val is not None:
            values.append({"value": str(val).strip(), "description": row[4].value})
    return {"kind": "dropdown", "values": values}


def extract_workbook(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    attrs = {}
    for sheet_name, label in SHEET_TO_LABEL.items():
        if sheet_name in wb.sheetnames:
            attrs[label] = extract_attribute_sheet(wb[sheet_name])
    return attrs


def build_folder_template():
    # "06_Transmittals/.../00_Starter Pack" holds one-off starter-pack resource
    # folders (a Revit template dir, "Schedule 13", "Shared Parameters") from
    # whichever project this template was last cloned from -- not a recurring
    # structural requirement, so we don't descend into it.
    STOP_AT = "00_Starter Pack"
    rel_folders = []
    for root, dirs, files in os.walk(FOLDER_TEMPLATE_ROOT):
        rel_root = os.path.relpath(root, FOLDER_TEMPLATE_ROOT).replace("\\", "/")
        if rel_root != "." and os.path.basename(rel_root) == STOP_AT:
            dirs[:] = []
        for d in dirs:
            full = os.path.join(root, d)
            rel = os.path.relpath(full, FOLDER_TEMPLATE_ROOT).replace("\\", "/")
            rel_folders.append(rel)
    return sorted(rel_folders)


def main():
    projects = {}
    issues = []

    for code in PROJECT_CODES:
        path, status, notes = find_canonical_workbook(code)
        if path is None:
            issues.append(f"{code}: UNRESOLVED -- no workbook found whose own 'Project' sheet says {code}. "
                           f"Candidates seen: {notes}. Excluded from naming_rules.json -- files for this "
                           f"project code will only be checked for folder structure/metadata, not naming pattern.")
            continue

        attrs = extract_workbook(path)
        description = None
        if "ProjectBuildingAsset" in attrs and attrs["ProjectBuildingAsset"]["kind"] == "dropdown":
            for v in attrs["ProjectBuildingAsset"]["values"]:
                if v["value"].startswith(code):
                    description = v["description"].split(" / ")[0].split(" - ", 1)[-1] if v["description"] else None
                    break

        contract_numbers = []
        if "Contract No" in attrs and attrs["Contract No"]["kind"] == "dropdown":
            contract_numbers = [v["value"] for v in attrs["Contract No"]["values"]]

        projects[code] = {
            "source_file": os.path.relpath(path, ONEDRIVE_ROOT),
            "description": description,
            "contract_numbers": contract_numbers,
            "attributes": attrs,
            "confidence": "high" if status == "ok" else "low_ambiguous_duplicate",
        }
        if status == "ambiguous_duplicates":
            issues.append(f"{code}: multiple internally-consistent workbooks found with near-identical sync "
                           f"timestamps (cannot reliably tell which is current) -- defaulted to "
                           f"{os.path.relpath(path, ONEDRIVE_ROOT)}, marked confidence=low. Others: {notes}. "
                           f"Please confirm manually, or prefer pulling the live Naming Standard from ACC once that's wired up.")
        elif notes:
            issues.append(f"{code}: picked {os.path.relpath(path, ONEDRIVE_ROOT)}; ignored mismatched file(s): {notes}")

    config = {
        "delimiter": "-",
        "naming_segment_order": NAMING_SEGMENT_ORDER,
        "required_segments": REQUIRED_SEGMENTS,
        "optional_trailing_segments": OPTIONAL_TRAILING_SEGMENTS,
        "metadata_required_fields": METADATA_REQUIRED_FIELDS,
        "folder_template": build_folder_template(),
        "projects": projects,
        "extraction_issues": issues,
    }

    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(config, f, indent=2, ensure_ascii=False)

    print(f"Wrote {os.path.abspath(OUT_PATH)}")
    print(f"Resolved {len(projects)}/{len(PROJECT_CODES)} project codes.")
    if issues:
        print("\nISSUES REQUIRING REVIEW:")
        for issue in issues:
            print(f"  - {issue}")


if __name__ == "__main__":
    main()
